from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from awscope.models import AwsResource, ScanResult

_HEADER_FILL = PatternFill("solid", fgColor="A8D8EA")
_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")

# Resource types that go into dedicated sheets instead of "All Resources"
_IAM_TYPES = {"iam:user", "iam:group", "iam:role", "iam:policy"}
_NET_TYPES = {
    "ec2:vpc", "ec2:subnet", "ec2:security-group", "ec2:internet-gateway",
    "ec2:nat-gateway", "ec2:vpc-endpoint", "ec2:route-table", "ec2:network-acl",
    "ec2:transit-gateway", "ec2:vpc-peering-connection",
    "elbv2:load-balancer", "elb:load-balancer",
    "acm:certificate",
}


def _header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autowidth(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 60)


def _tags_str(tags: dict) -> str:
    return ", ".join(f"{k}={v}" for k, v in sorted(tags.items()))


def _fmt_list(items: list) -> str:
    if not items:
        return ""
    # Flatten — each item may itself be a list (e.g. multiple trust principals)
    flat = []
    for item in items:
        if isinstance(item, list):
            flat.extend(str(i) for i in item)
        else:
            flat.append(str(item))
    return "\n".join(flat)


# ── Sheet 1: Summary ───────────────────────────────────────────────────────

def _write_summary(ws, result: ScanResult) -> None:
    _header(ws, ["Group", "Resource Count", "Resource Types", "Accounts", "Regions"])
    for group in sorted(result.groups, key=lambda g: g.group_name):
        types = sorted({r.resource_type for r in group.resources})
        accounts = sorted({r.account_alias for r in group.resources})
        regions = sorted({r.region for r in group.resources})
        row = ws.append([
            group.group_name,
            len(group.resources),
            ", ".join(types),
            ", ".join(accounts),
            ", ".join(regions),
        ])
    _autowidth(ws)


# ── Sheet 2: All Resources ─────────────────────────────────────────────────

def _write_all_resources(ws, result: ScanResult) -> None:
    _header(ws, ["Group", "Name", "Resource Type", "Resource ID", "ARN",
                 "Region", "Account Alias", "Account ID", "Status", "Tags"])
    rows = []
    for group in result.groups:
        for r in group.resources:
            if r.resource_type in _IAM_TYPES or r.resource_type in _NET_TYPES:
                continue
            rows.append((group.group_name, r))
    rows.sort(key=lambda x: (x[0], x[1].resource_type, x[1].name))
    for group_name, r in rows:
        ws.append([
            group_name, r.name, r.resource_type, r.resource_id,
            r.arn, r.region, r.account_alias, r.account_id,
            r.status, _tags_str(r.tags),
        ])
    _autowidth(ws)


def _extract_trust_principals(policy_doc: dict) -> list[str]:
    """Flatten all Principal values from an AssumeRolePolicyDocument into strings."""
    principals = []
    for statement in policy_doc.get("Statement", []):
        principal = statement.get("Principal", {})
        if isinstance(principal, str):
            principals.append(principal)
        elif isinstance(principal, dict):
            for v in principal.values():
                if isinstance(v, list):
                    principals.extend(str(i) for i in v)
                else:
                    principals.append(str(v))
    return principals


# ── Sheet 3: IAM Audit ─────────────────────────────────────────────────────

def _write_iam(ws, result: ScanResult) -> None:
    all_iam = [r for r in result.resources if r.resource_type in _IAM_TYPES]

    def _section(title: str, columns: list[str], resource_type: str, row_fn):
        ws.append([title])
        ws[ws.max_row][0].font = Font(bold=True, size=12)
        _header_row = ws.max_row + 1
        ws.append(columns)
        for cell in ws[ws.max_row]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for r in sorted(all_iam, key=lambda x: x.name):
            if r.resource_type == resource_type:
                ws.append(row_fn(r))
        ws.append([])  # blank separator

    _section(
        "IAM Users", ["Account", "Username", "ARN", "MFA Enabled", "Groups",
                       "Attached Policies", "Inline Policies", "Tags"],
        "iam:user",
        lambda r: [
            r.account_alias, r.name, r.arn,
            str(r.raw.get("mfa_enabled", False)),
            _fmt_list(r.raw.get("groups", [])),
            _fmt_list(r.raw.get("attached_policies", [])),
            _fmt_list(r.raw.get("inline_policies", [])),
            _tags_str(r.tags),
        ],
    )

    _section(
        "IAM Roles", ["Account", "Role Name", "ARN", "Trust Principals",
                       "Attached Policies", "Inline Policies"],
        "iam:role",
        lambda r: [
            r.account_alias, r.name, r.arn,
            _fmt_list(_extract_trust_principals(r.raw.get("AssumeRolePolicyDocument", {}))),
            _fmt_list(r.raw.get("attached_policies", [])),
            _fmt_list(r.raw.get("inline_policies", [])),
        ],
    )

    _section(
        "IAM Groups", ["Account", "Group Name", "ARN", "Members",
                        "Attached Policies", "Inline Policies"],
        "iam:group",
        lambda r: [
            r.account_alias, r.name, r.arn,
            _fmt_list(r.raw.get("members", [])),
            _fmt_list(r.raw.get("attached_policies", [])),
            _fmt_list(r.raw.get("inline_policies", [])),
        ],
    )

    _section(
        "Customer-Managed Policies", ["Account", "Policy Name", "ARN", "Description"],
        "iam:policy",
        lambda r: [
            r.account_alias, r.name, r.arn,
            r.raw.get("Description", ""),
        ],
    )

    # Apply wrap text to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = _WRAP
    _autowidth(ws)


# ── Sheet 4: Networking ────────────────────────────────────────────────────

def _write_networking(ws, result: ScanResult) -> None:
    net_resources = [r for r in result.resources if r.resource_type in _NET_TYPES]

    def _section(title: str, columns: list[str], rtype: str, row_fn):
        matching = [r for r in net_resources if r.resource_type == rtype]
        if not matching:
            return
        ws.append([title])
        ws[ws.max_row][0].font = Font(bold=True, size=12)
        ws.append(columns)
        for cell in ws[ws.max_row]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for r in sorted(matching, key=lambda x: (x.account_alias, x.region, x.name)):
            ws.append(row_fn(r))
        ws.append([])

    _section(
        "VPCs", ["Account", "Region", "VPC ID", "Name", "CIDR", "Default", "State", "Tags"],
        "ec2:vpc",
        lambda r: [
            r.account_alias, r.region, r.resource_id, r.name,
            r.raw.get("CidrBlock", ""),
            str(r.raw.get("IsDefault", False)),
            r.status, _tags_str(r.tags),
        ],
    )

    _section(
        "Security Groups",
        ["Account", "Region", "Group ID", "Name", "VPC ID", "Description",
         "Inbound Rules", "Outbound Rules", "Tags"],
        "ec2:security-group",
        lambda r: [
            r.account_alias, r.region, r.resource_id, r.name,
            r.raw.get("VpcId", ""),
            r.raw.get("Description", ""),
            _fmt_list([
                f"{p.get('IpProtocol')}:{p.get('FromPort', '*')}-{p.get('ToPort', '*')}:"
                + ",".join(ip.get("CidrIp", ip.get("CidrIpv6", "")) for ip in p.get("IpRanges", []) + p.get("Ipv6Ranges", []))
                for p in r.raw.get("IpPermissions", [])
            ]),
            _fmt_list([
                f"{p.get('IpProtocol')}:{p.get('FromPort', '*')}-{p.get('ToPort', '*')}:"
                + ",".join(ip.get("CidrIp", ip.get("CidrIpv6", "")) for ip in p.get("IpRanges", []) + p.get("Ipv6Ranges", []))
                for p in r.raw.get("IpPermissionsEgress", [])
            ]),
            _tags_str(r.tags),
        ],
    )

    _section(
        "Subnets",
        ["Account", "Region", "Subnet ID", "Name", "VPC ID", "CIDR", "AZ", "Public", "Tags"],
        "ec2:subnet",
        lambda r: [
            r.account_alias, r.region, r.resource_id, r.name,
            r.raw.get("VpcId", ""),
            r.raw.get("CidrBlock", ""),
            r.raw.get("AvailabilityZone", ""),
            str(r.raw.get("MapPublicIpOnLaunch", False)),
            _tags_str(r.tags),
        ],
    )

    _section(
        "Load Balancers",
        ["Account", "Region", "Name", "Type", "Scheme", "DNS Name", "VPC ID", "State", "Tags"],
        "elbv2:load-balancer",
        lambda r: [
            r.account_alias, r.region, r.name,
            r.raw.get("Type", ""),
            r.raw.get("Scheme", ""),
            r.raw.get("DNSName", ""),
            r.raw.get("VpcId", ""),
            r.status, _tags_str(r.tags),
        ],
    )

    _section(
        "Classic Load Balancers",
        ["Account", "Region", "Name", "Scheme", "DNS Name", "VPC ID"],
        "elb:load-balancer",
        lambda r: [
            r.account_alias, r.region, r.name,
            r.raw.get("Scheme", ""),
            r.raw.get("DNSName", ""),
            r.raw.get("VPCId", ""),
        ],
    )

    _section(
        "Gateways & Endpoints",
        ["Account", "Region", "Type", "ID", "Name", "State", "Tags"],
        "ec2:internet-gateway",
        lambda r: [r.account_alias, r.region, r.resource_type, r.resource_id, r.name, r.status, _tags_str(r.tags)],
    )
    for rtype in ("ec2:nat-gateway", "ec2:vpc-endpoint", "ec2:transit-gateway"):
        _section(
            "", ["Account", "Region", "Type", "ID", "Name", "State", "Tags"],
            rtype,
            lambda r: [r.account_alias, r.region, r.resource_type, r.resource_id, r.name, r.status, _tags_str(r.tags)],
        )

    _section(
        "ACM Certificates",
        ["Account", "Region", "Domain", "ARN", "Status"],
        "acm:certificate",
        lambda r: [r.account_alias, r.region, r.name, r.arn, r.status],
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = _WRAP
    _autowidth(ws)


# ── public API ─────────────────────────────────────────────────────────────

def build_excel(result: ScanResult) -> BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    _write_summary(ws1, result)

    ws2 = wb.create_sheet("All Resources")
    _write_all_resources(ws2, result)

    ws3 = wb.create_sheet("IAM Audit")
    _write_iam(ws3, result)

    ws4 = wb.create_sheet("Networking")
    _write_networking(ws4, result)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
