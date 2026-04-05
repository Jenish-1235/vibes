from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from awscope.exporter import _extract_trust_principals, _fmt_list, build_excel
from awscope.models import AwsResource, ResourceGroup, ScanResult


def _resource(rid, name, rtype, region="us-east-1", tags=None, raw=None):
    return AwsResource(
        resource_id=rid, name=name, resource_type=rtype, arn=f"arn::{rid}",
        region=region, account_alias="personal", account_id="123",
        status="active", tags=tags or {}, raw=raw or {},
    )


def _make_result():
    resources = [
        _resource("i-123", "hagrid-api", "ec2:instance"),
        _resource("my-bucket", "my-bucket", "s3:bucket", region="global"),
        _resource("user-alice", "alice", "iam:user", raw={
            "mfa_enabled": True,
            "attached_policies": ["ReadOnlyAccess"],
            "inline_policies": [],
            "groups": ["developers"],
        }),
        _resource("role-deployer", "deployer", "iam:role", raw={
            "attached_policies": ["AmazonEC2FullAccess"],
            "inline_policies": [],
            "AssumeRolePolicyDocument": {
                "Statement": [{
                    "Principal": {"Service": "ec2.amazonaws.com"},
                    "Effect": "Allow",
                    "Action": "sts:AssumeRole",
                }]
            },
        }),
        _resource("vpc-abc", "vpc-abc", "ec2:vpc", raw={"CidrBlock": "10.0.0.0/16", "IsDefault": False}),
        _resource("sg-xyz", "my-sg", "ec2:security-group", tags={"Name": "my-sg"}, raw={
            "VpcId": "vpc-abc", "Description": "test sg",
            "IpPermissions": [], "IpPermissionsEgress": [],
        }),
    ]
    groups = [
        ResourceGroup(group_name="hagrid", resources=resources[:2]),
        ResourceGroup(group_name="miscellaneous", resources=resources[2:]),
    ]
    return ScanResult(
        scanned_at="2026-04-04T10:00:00Z",
        accounts=["personal"],
        resources=resources,
        groups=groups,
    )


# ── _fmt_list ──────────────────────────────────────────────────────────────

def test_fmt_list_empty():
    assert _fmt_list([]) == ""


def test_fmt_list_strings():
    assert _fmt_list(["a", "b"]) == "a\nb"


def test_fmt_list_nested_list():
    # Principal values can be a list of strings
    assert _fmt_list([["arn:aws:iam::123:root", "arn:aws:iam::456:root"]]) == \
        "arn:aws:iam::123:root\narn:aws:iam::456:root"


def test_fmt_list_mixed():
    assert _fmt_list(["string", ["list-item"]]) == "string\nlist-item"


# ── _extract_trust_principals ──────────────────────────────────────────────

def test_extract_trust_principals_service():
    doc = {"Statement": [{"Principal": {"Service": "ec2.amazonaws.com"}}]}
    assert _extract_trust_principals(doc) == ["ec2.amazonaws.com"]


def test_extract_trust_principals_aws_string():
    doc = {"Statement": [{"Principal": {"AWS": "arn:aws:iam::123:root"}}]}
    assert _extract_trust_principals(doc) == ["arn:aws:iam::123:root"]


def test_extract_trust_principals_aws_list():
    # Real AWS responses often return a list here
    doc = {"Statement": [{"Principal": {"AWS": ["arn:aws:iam::123:root", "arn:aws:iam::456:root"]}}]}
    result = _extract_trust_principals(doc)
    assert "arn:aws:iam::123:root" in result
    assert "arn:aws:iam::456:root" in result


def test_extract_trust_principals_star():
    doc = {"Statement": [{"Principal": "*"}]}
    assert _extract_trust_principals(doc) == ["*"]


def test_extract_trust_principals_empty():
    assert _extract_trust_principals({}) == []


def test_extract_trust_principals_multiple_statements():
    doc = {"Statement": [
        {"Principal": {"Service": "lambda.amazonaws.com"}},
        {"Principal": {"AWS": "arn:aws:iam::123:role/MyRole"}},
    ]}
    result = _extract_trust_principals(doc)
    assert len(result) == 2


# ── build_excel ────────────────────────────────────────────────────────────

def test_build_excel_returns_bytes():
    result = _make_result()
    buf = build_excel(result)
    assert isinstance(buf, BytesIO)
    assert buf.tell() == 0


def test_excel_has_four_sheets():
    result = _make_result()
    buf = build_excel(result)
    wb = load_workbook(buf)
    assert set(wb.sheetnames) == {"Summary", "All Resources", "IAM Audit", "Networking"}


def test_summary_sheet_has_groups():
    result = _make_result()
    buf = build_excel(result)
    wb = load_workbook(buf)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    group_names = [r[0] for r in rows[1:] if r[0]]
    assert "hagrid" in group_names
    assert "miscellaneous" in group_names


def test_all_resources_excludes_iam_and_net():
    result = _make_result()
    buf = build_excel(result)
    wb = load_workbook(buf)
    ws = wb["All Resources"]
    rows = list(ws.iter_rows(values_only=True))
    types_in_sheet = {r[2] for r in rows[1:] if r[2]}
    assert "iam:user" not in types_in_sheet
    assert "iam:role" not in types_in_sheet
    assert "ec2:vpc" not in types_in_sheet
    assert "ec2:instance" in types_in_sheet


def test_iam_sheet_has_users_and_roles():
    result = _make_result()
    buf = build_excel(result)
    wb = load_workbook(buf)
    ws = wb["IAM Audit"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("IAM Users" in v for v in all_values)
    assert any("IAM Roles" in v for v in all_values)


def test_iam_role_trust_principal_service():
    # Role with Service principal — must not raise
    result = _make_result()
    buf = build_excel(result)
    wb = load_workbook(buf)
    ws = wb["IAM Audit"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("ec2.amazonaws.com" in v for v in all_values)


def test_iam_role_with_list_principal():
    # Real-world: Principal.AWS is a list — must not raise
    r = _resource("role-multi", "multi-principal-role", "iam:role", raw={
        "attached_policies": [],
        "inline_policies": [],
        "AssumeRolePolicyDocument": {
            "Statement": [{
                "Principal": {"AWS": ["arn:aws:iam::111:root", "arn:aws:iam::222:root"]},
                "Effect": "Allow",
                "Action": "sts:AssumeRole",
            }]
        },
    })
    result = ScanResult(
        scanned_at="2026-04-04T10:00:00Z", accounts=["personal"],
        resources=[r],
        groups=[ResourceGroup(group_name="misc", resources=[r])],
    )
    buf = build_excel(result)  # must not raise
    assert isinstance(buf, BytesIO)


def test_networking_sheet_has_vpc_and_sg():
    result = _make_result()
    buf = build_excel(result)
    wb = load_workbook(buf)
    ws = wb["Networking"]
    all_values = [str(c.value or "") for row in ws.iter_rows() for c in row]
    assert any("VPC" in v for v in all_values)
    assert any("Security Group" in v for v in all_values)


def test_raw_with_datetime_does_not_crash_cache():
    # Regression: boto3 responses contain datetime objects in raw
    from awscope.cache import _scan_result_to_dict, write_cache
    import json, tempfile, os
    r = _resource("i-dt", "dt-instance", "ec2:instance", raw={
        "LaunchTime": datetime(2026, 4, 4, 10, 0, 0),
        "StateTransitionReason": "User initiated",
    })
    result = ScanResult(
        scanned_at="2026-04-04T10:00:00Z", accounts=["personal"],
        resources=[r],
        groups=[ResourceGroup(group_name="misc", resources=[r])],
    )
    d = _scan_result_to_dict(result)
    # Must serialize without TypeError
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        path = f.name
    try:
        os.environ["AWSCOPE_CACHE_PATH"] = path
        write_cache(result)
        with open(path) as f:
            loaded = json.load(f)
        assert loaded["resources"][0]["raw"]["LaunchTime"] == "2026-04-04T10:00:00"
    finally:
        os.unlink(path)
        os.environ.pop("AWSCOPE_CACHE_PATH", None)
